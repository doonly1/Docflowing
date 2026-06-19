"""文件库文件操作 - 文件内容管理与下载"""

import os
import io
import re
import shutil
from flask import Blueprint, request, jsonify, send_file, g

from server.auth import login_required
from fb.database import get_db
from fb.decorators import _require_fb_permission, require_fb_perm, _ensure_local_fb_route, _get_node_identity, require_not_locked
from fb.routes_files import _trigger_fb_sync
from server.workspace import _get_workspace_dir


def _is_path_safe(base_dir: str, target_path: str) -> bool:
    """检查 target_path 是否在 base_dir 内部，防止路径穿越攻击。

    使用 os.path.realpath 解析符号链接后比对，避免 .. 穿越。
    """
    try:
        base = os.path.realpath(base_dir)
        target = os.path.realpath(target_path)
        return target.startswith(base + os.sep) or target == base
    except OSError:
        return False


def _validate_save_dir(dest_dir: str) -> tuple[bool, str]:
    """验证另存目标目录是否安全（必须在用户工作空间内）。

    Returns:
        (is_safe, error_message)
    """
    if not dest_dir:
        return False, '未指定保存路径'
    try:
        dest_dir = os.path.abspath(dest_dir)
    except Exception:
        return False, '保存路径无效'
    allowed_base = os.path.realpath(_get_workspace_dir())
    if not _is_path_safe(allowed_base, dest_dir):
        return False, '保存路径不能在工作空间之外'
    return True, ''


def _unique_path(dest_dir, name):
    """如果目标路径已存在，自动添加数字后缀以避免覆盖（有上限）"""
    target = os.path.join(dest_dir, name)
    if not os.path.exists(target):
        return target
    base, ext = os.path.splitext(name)
    for n in range(1, 10000):
        new_name = f"{base} ({n}){ext}"
        target = os.path.join(dest_dir, new_name)
        if not os.path.exists(target):
            return target
    # 达到上限，用时间戳兜底
    import time
    return os.path.join(dest_dir, f"{base}_{int(time.time() * 1000)}{ext}")

fb_bp = Blueprint('fb', __name__, url_prefix='/api/fb')


@fb_bp.route('/<fb_id>/local-files/content', methods=['PUT'])
@login_required
@require_fb_perm('edit')
@require_not_locked
@_ensure_local_fb_route
def save_local_file_content(filebase_id):
    """保存文件内容"""
    if getattr(g, 'is_remote_fb', False):
        from p2p import proxy as p2p_proxy
        node = _get_node_identity()
        info = g.remote_fb_info
        data = request.get_json() or {}
        result = p2p_proxy.remote_save_file(
            info['owner_addr'], node, filebase_id,
            (data.get('path') or '').strip(),
            data.get('content', ''),
            data.get('client_mtime', 0)
        )
        return jsonify(result or {'success': False, 'message': '远程节点不可用'})

    db = get_db()
    data = request.get_json() or {}
    path = (data.get('path') or '').strip()
    content = data.get('content', '')
    client_mtime = data.get('client_mtime', 0)

    if not path:
        return jsonify({'success': False, 'message': '未指定文件路径'})

    local_path, target = _resolve_local_path(db, filebase_id, path)
    if local_path is None:
        return jsonify({'success': False, 'message': '文件库不存在或路径非法'})

    if os.path.isfile(target) and client_mtime:
        current_mtime = os.stat(target).st_mtime
        if current_mtime != client_mtime:
            return jsonify({
                'success': False,
                'conflict': True,
                'message': '文件已被他人修改，是否覆盖？',
                'server_mtime': current_mtime
            }), 409

    os.makedirs(os.path.dirname(target), exist_ok=True)
    with open(target, 'w', encoding='utf-8') as f:
        f.write(content)
    stat = os.stat(target)
    _trigger_fb_sync(filebase_id)
    return jsonify({'success': True, 'mtime': stat.st_mtime})


@fb_bp.route('/<fb_id>/local-files/content', methods=['GET'])
@login_required
@require_fb_perm('view')
@_ensure_local_fb_route
def get_local_file_content(filebase_id):
    """获取文件内容"""
    if getattr(g, 'is_remote_fb', False):
        from p2p import proxy as p2p_proxy
        node = _get_node_identity()
        info = g.remote_fb_info
        result = p2p_proxy.remote_get_file_content(info['owner_addr'], node, filebase_id, request.args.get('path', ''))
        return jsonify(result or {'success': False, 'message': '远程节点不可用'})

    db = get_db()
    kb_row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (filebase_id,)).fetchone()
    if not kb_row:
        return jsonify({'success': False, 'message': '文件库不存在'})

    local_path = kb_row['local_path']
    rel_path = request.args.get('path', '').strip()
    if not rel_path:
        return jsonify({'success': False, 'message': '未指定文件路径'})

    file_path = os.path.normpath(os.path.join(local_path, rel_path))
    if not file_path.startswith(os.path.normpath(local_path) + os.sep) and file_path != os.path.normpath(local_path):
        return jsonify({'success': False, 'message': '不允许访问上级目录'})

    if not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': '文件不存在'})

    ext = os.path.splitext(file_path)[1].lower()

    if ext in ('.md', '.txt', '.html', '.htm', '.xml', '.json', '.csv', '.yaml', '.yml', '.py', '.js', '.css'):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return jsonify({'success': True, 'content': content, 'file_type': ext})
        except UnicodeDecodeError:
            return jsonify({'success': False, 'message': '无法以文本方式读取此文件'})

    if ext == '.docx':
        try:
            from docx import Document
            doc = Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs]
            return jsonify({'success': True, 'content': '\n'.join(paragraphs), 'file_type': ext})
        except Exception:
            return jsonify({'success': False, 'message': '无法读取 docx 内容'})

    if ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f'# {sheet_name}')
                for row in ws.iter_rows(values_only=True):
                    lines.append('\t'.join(str(c) if c is not None else '' for c in row))
            return jsonify({'success': True, 'content': '\n'.join(lines), 'file_type': ext})
        except Exception:
            return jsonify({'success': False, 'message': '无法读取 xlsx 内容'})

    return jsonify({'success': False, 'message': '不支持在线查看此文件类型的内容'})


SUPPORTED_PREVIEW_EXTS = {'.docx', '.pptx', '.ppt', '.xlsx', '.xls'}


@fb_bp.route('/<fb_id>/local-files/preview', methods=['GET'])
@login_required
@require_fb_perm('view')
@_ensure_local_fb_route
def file_preview(filebase_id):
    """预览文件"""
    if getattr(g, 'is_remote_fb', False):
        from p2p import proxy as p2p_proxy
        node = _get_node_identity()
        info = g.remote_fb_info
        result = p2p_proxy.remote_get_file_content(info['owner_addr'], node, filebase_id, request.args.get('path', ''))
        return jsonify(result or {'success': False, 'message': '远程节点不可用'})

    db = get_db()
    kb_row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (filebase_id,)).fetchone()
    if not kb_row:
        return jsonify({'success': False, 'message': '文件库不存在'})

    local_path = kb_row['local_path']
    rel_path = request.args.get('path', '').strip()
    if not rel_path:
        return jsonify({'success': False, 'message': '未指定文件路径'})

    file_path = os.path.normpath(os.path.join(local_path, rel_path))
    if not file_path.startswith(os.path.normpath(local_path) + os.sep) and file_path != os.path.normpath(local_path):
        return jsonify({'success': False, 'message': '不允许访问上级目录'})

    if not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': '文件不存在'})

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_PREVIEW_EXTS:
        return jsonify({'success': False, 'message': f'不支持的预览格式: {ext}'})

    try:
        from kb.sync_converters import MarkItDownConverter
        converter = MarkItDownConverter()
        markdown = converter.convert(file_path)
        if markdown is None:
            return jsonify({'success': False, 'message': '文件转换失败'})
        return jsonify({'success': True, 'markdown': markdown, 'file_type': ext})
    except Exception as e:
        return jsonify({'success': False, 'message': f'预览失败: {str(e)}'})


@fb_bp.route('/<fb_id>/local-files/download', methods=['GET'])
@login_required
@require_fb_perm('view')
@_ensure_local_fb_route
def download_local_file(filebase_id):
    """下载文件"""
    if getattr(g, 'is_remote_fb', False):
        from p2p import proxy as p2p_proxy
        node = _get_node_identity()
        info = g.remote_fb_info
        resp = p2p_proxy.remote_download_file(info['owner_addr'], node, filebase_id, request.args.get('path', ''))
        if resp:
            from flask import Response
            return Response(resp.iter_content(chunk_size=8192), content_type=resp.headers.get('Content-Type', 'application/octet-stream'))
        return jsonify({'success': False, 'message': '远程节点不可用'})

    db = get_db()
    kb_row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (filebase_id,)).fetchone()
    if not kb_row:
        return jsonify({'success': False, 'message': '文件库不存在'})

    local_path = kb_row['local_path']
    rel_path = request.args.get('path', '').strip()
    if not rel_path:
        return jsonify({'success': False, 'message': '未指定文件路径'})

    file_path = os.path.normpath(os.path.join(local_path, rel_path))
    if not file_path.startswith(os.path.normpath(local_path) + os.sep) and file_path != os.path.normpath(local_path):
        return jsonify({'success': False, 'message': '不允许访问上级目录'})

    if not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': '文件不存在'})

    return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))


@fb_bp.route('/<fb_id>/local-files/batch-download', methods=['POST'])
@login_required
@require_fb_perm('view')
def batch_download_local(filebase_id):
    """批量下载文件"""
    db = get_db()
    kb_row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (filebase_id,)).fetchone()
    if not kb_row:
        return jsonify({'success': False, 'message': '文件库不存在'})

    local_path = kb_row['local_path']
    data = request.get_json() or {}
    paths = data.get('paths', [])

    if not paths:
        return jsonify({'success': False, 'message': '请选择文件'})

    import zipfile
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for rel_path in paths:
            abs_path = os.path.normpath(os.path.join(local_path, rel_path))
            if not abs_path.startswith(os.path.normpath(local_path) + os.sep) and abs_path != os.path.normpath(local_path):
                continue
            if os.path.isfile(abs_path):
                zf.write(abs_path, os.path.basename(abs_path))
            elif os.path.isdir(abs_path):
                for root, dirs, files in os.walk(abs_path):
                    for f in files:
                        fp = os.path.join(root, f)
                        arc = os.path.relpath(fp, os.path.dirname(abs_path)).replace('\\', '/')
                        zf.write(fp, arc)

    memory_file.seek(0)
    return send_file(memory_file, mimetype='application/zip',
                     as_attachment=True, download_name='files.zip')


@fb_bp.route('/<fb_id>/local-files/save-as', methods=['POST'])
@login_required
@require_fb_perm('view')
def save_local_file_as(filebase_id):
    """另存文件"""
    data = request.get_json() or {}
    rel_path = data.get('path', '').strip()
    save_path = data.get('save_path', '').strip()

    if not rel_path:
        return jsonify({'success': False, 'message': '未指定源文件路径'})
    if not save_path:
        return jsonify({'success': False, 'message': '未指定保存路径'})

    # 校验保存路径：必须位于用户工作空间内（本地和远程分支共用）
    safe, err = _validate_save_dir(os.path.dirname(save_path))
    if not safe:
        return jsonify({'success': False, 'message': err})

    if getattr(g, 'is_remote_fb', False):
        from p2p import proxy as p2p_proxy
        node = _get_node_identity()
        info = g.remote_fb_info
        try:
            resp = p2p_proxy.remote_download_file(info['owner_addr'], node, filebase_id, rel_path)
            if not resp:
                return jsonify({'success': False, 'message': '远程节点不可用'})
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            with open(save_path, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            return jsonify({'success': True, 'message': '文件已保存'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500

    db = get_db()
    kb_row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (filebase_id,)).fetchone()
    if not kb_row:
        return jsonify({'success': False, 'message': '文件库不存在'})

    local_path = kb_row['local_path']
    rel_path = data.get('path', '').strip()

    src_file = os.path.normpath(os.path.join(local_path, rel_path))
    if not src_file.startswith(os.path.normpath(local_path) + os.sep) and src_file != os.path.normpath(local_path):
        return jsonify({'success': False, 'message': '路径非法'})
    if not os.path.isfile(src_file):
        return jsonify({'success': False, 'message': '源文件不存在'})

    try:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        shutil.copy2(src_file, save_path)
        return jsonify({'success': True, 'message': '文件已保存'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500


@fb_bp.route('/<fb_id>/local-files/batch-save-as', methods=['POST'])
@login_required
@require_fb_perm('view')
def batch_save_local_files(filebase_id):
    """批量另存文件"""
    data = request.get_json() or {}
    paths = data.get('paths', [])
    dest_dir = data.get('dest_dir', '').strip()

    if not paths:
        return jsonify({'success': False, 'message': '请选择文件'})
    if not dest_dir:
        return jsonify({'success': False, 'message': '未指定目标目录'})

    # 校验目标目录：必须位于用户工作空间内
    safe, err = _validate_save_dir(dest_dir)
    if not safe:
        return jsonify({'success': False, 'message': err})

    if getattr(g, 'is_remote_fb', False):
        from p2p import proxy as p2p_proxy
        node = _get_node_identity()
        info = g.remote_fb_info
        try:
            os.makedirs(dest_dir, exist_ok=True)
            for rel_path in paths:
                if rel_path.endswith('/') or rel_path.endswith('\\'):
                    continue
                fname = os.path.basename(rel_path.replace('\\', '/'))
                save_path = _unique_path(dest_dir, fname)
                resp = p2p_proxy.remote_download_file(info['owner_addr'], node, filebase_id, rel_path)
                if not resp:
                    continue
                with open(save_path, 'wb') as f:
                    for chunk in resp.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
            return jsonify({'success': True, 'message': '文件已保存'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500

    db = get_db()
    kb_row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (filebase_id,)).fetchone()
    if not kb_row:
        return jsonify({'success': False, 'message': '文件库不存在'})

    local_path = kb_row['local_path']
    try:
        os.makedirs(dest_dir, exist_ok=True)
        for rel_path in paths:
            abs_path = os.path.normpath(os.path.join(local_path, rel_path))
            if not abs_path.startswith(os.path.normpath(local_path) + os.sep) and abs_path != os.path.normpath(local_path):
                continue
            fname = os.path.basename(rel_path.replace('\\', '/'))
            target = _unique_path(dest_dir, fname)
            if os.path.isfile(abs_path):
                shutil.copy2(abs_path, target)
            elif os.path.isdir(abs_path):
                shutil.copytree(abs_path, target, dirs_exist_ok=True)
        return jsonify({'success': True, 'message': '文件已保存'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500


@fb_bp.route('/<fb_id>/local-files/open', methods=['GET'])
@login_required
@require_fb_perm('view')
def open_local_file(filebase_id):
    """打开文件"""
    db = get_db()
    kb_row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (filebase_id,)).fetchone()
    if not kb_row:
        return jsonify({'success': False, 'message': '文件库不存在'})

    local_path = kb_row['local_path']
    rel_path = request.args.get('path', '').strip()
    if not rel_path:
        return jsonify({'success': False, 'message': '未指定文件路径'})

    file_path = os.path.normpath(os.path.join(local_path, rel_path))
    if not file_path.startswith(os.path.normpath(local_path) + os.sep) and file_path != os.path.normpath(local_path):
        return jsonify({'success': False, 'message': '不允许访问上级目录'})

    if not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': '文件不存在'})

    ext = os.path.splitext(file_path)[1].lower()
    previewable_exts = {'.pdf', '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg', '.txt', '.csv'}
    as_attachment = ext not in previewable_exts

    mimetype = None
    if ext == '.pdf':
        mimetype = 'application/pdf'
    return send_file(file_path, as_attachment=as_attachment, download_name=os.path.basename(file_path), mimetype=mimetype)


@fb_bp.route('/<fb_id>/local-files/open-with-app', methods=['GET'])
@login_required
@require_fb_perm('view')
def open_with_app(filebase_id):
    """用系统默认软件打开文件"""
    import platform

    db = get_db()
    kb_row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (filebase_id,)).fetchone()
    if not kb_row:
        return jsonify({'success': False, 'message': '文件库不存在'})

    local_path = kb_row['local_path']
    rel_path = request.args.get('path', '').strip()
    if not rel_path:
        return jsonify({'success': False, 'message': '未指定文件路径'})

    file_path = os.path.normpath(os.path.join(local_path, rel_path))
    if not file_path.startswith(os.path.normpath(local_path) + os.sep) and file_path != os.path.normpath(local_path):
        return jsonify({'success': False, 'message': '路径非法'})

    if not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': '文件不存在'})

    try:
        system = platform.system()
        if system == 'Windows':
            os.startfile(file_path)
        elif system == 'Darwin':
            import subprocess
            subprocess.run(['open', file_path], check=True)
        else:
            import subprocess
            subprocess.run(['xdg-open', file_path], check=True)
        return jsonify({'success': True, 'message': '已用本地软件打开'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'打开失败: {str(e)}'}), 500
