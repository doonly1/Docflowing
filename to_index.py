# -*- coding: utf-8 -*-
"""
构建索引工具
整理目录文件到Excel，生成可跳转的文件索引
"""

import os
from datetime import datetime
from doc_process import doc_to_docx
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def format_file_size(size_bytes):
    if size_bytes == 0:
        return "0 B"
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    unit_index = 0
    size = float(size_bytes)
    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1
    if unit_index == 0:
        return f"{int(size)} B"
    elif size < 10:
        return f"{size:.2f} {units[unit_index]}"
    elif size < 100:
        return f"{size:.1f} {units[unit_index]}"
    else:
        return f"{int(size)} {units[unit_index]}"


def collect_file_info(root_dir):
    file_info_list = []
    root_dir = os.path.abspath(root_dir)
    for root, dirs, files in os.walk(root_dir):
        for file in files:
            file_path = os.path.join(root, file)
            rel_path = os.path.relpath(file_path, root_dir)
            try:
                stat_info = os.stat(file_path)
                mod_time = datetime.fromtimestamp(stat_info.st_mtime).strftime('%Y-%m-%d')
                file_size = stat_info.st_size
                formatted_size = format_file_size(file_size)
                path_parts = rel_path.split(os.sep)
                filename = path_parts[-1]
                dir_levels = path_parts[:-1]
                file_info = {
                    'filename': filename,
                    'rel_path': rel_path,
                    'abs_path': file_path,
                    'dir_levels': dir_levels,
                    'mod_time': mod_time,
                    'file_size': formatted_size
                }
                file_info_list.append(file_info)
            except (PermissionError, FileNotFoundError) as e:
                print(f"无法访问文件 {file_path}: {e}")
                continue
    return file_info_list


def create_excel(file_info_list, root_dir, output_file='file_list.xlsx'):
    if not file_info_list:
        print("没有找到任何文件")
        return
    max_depth = max(len(info['dir_levels']) for info in file_info_list)
    
    wb = Workbook()
    ws = wb.active
    ws.title = '文件列表'
    
    # 写入表头
    columns = ['序号', '文件名']
    for i in range(max_depth):
        columns.append(f'目录层{i+1}')
    columns.extend(['修改日期', '文件大小'])
    
    for col, header in enumerate(columns, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入数据
    for row_idx, info in enumerate(file_info_list, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)  # 序号
        # 超链接（文件名可点击）
        cell = ws.cell(row=row_idx, column=2, value=info['filename'])
        cell.hyperlink = info['rel_path']
        cell.font = Font()
        # 目录层级
        for col_idx, level in enumerate(info['dir_levels'], start=3):
            ws.cell(row=row_idx, column=col_idx, value=level)
        # 修改日期和文件大小
        ws.cell(row=row_idx, column=len(columns)-1, value=info['mod_time'])
        ws.cell(row=row_idx, column=len(columns), value=info['file_size'])
    
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value or '')) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output_path = os.path.join(root_dir, output_file)
    wb.save(output_path)
    print(f"Excel文件已生成: {output_path}")
    print(f"共整理了 {len(file_info_list)} 个文件")


def build_index(workdir):
    try:
        doc_to_docx(workdir)
    except Exception as e:
        print(f"  docx转换失败: {e}")
    print(f"正在扫描目录: {workdir}")
    file_info_list = collect_file_info(workdir)
    create_excel(file_info_list, workdir, 'file_index.xlsx')


def build_index_from_metadata(metadata_list, folder_name, output_dir):
    """从前端传入的文件元信息直接构建索引（无需读取服务器文件系统）

    Args:
        metadata_list: 前端 FileList 提取的元信息列表
            [{name, path, size, lastModified}, ...]
        folder_name: 文件夹名称（用于 Excel 中超链接的相对路径前缀）
        output_dir: 输出目录（服务器临时目录）

    Returns:
        output_path: 生成的 Excel 文件路径
    """
    file_info_list = []
    for item in metadata_list:
        rel_path = item.get('path', item.get('name', ''))
        filename = item.get('name', rel_path.split('/')[-1] if '/' in rel_path else rel_path)
        # webkitRelativePath 格式: folderName/sub/file.docx
        # 去掉第一层 folderName，使目录层级与本地模式一致
        path_parts = rel_path.split('/') if rel_path else [filename]
        if len(path_parts) > 1 and path_parts[0] == folder_name:
            path_parts = path_parts[1:]
        # 构建不含根目录名的相对路径（与本地 collect_file_info 一致）
        rel_path_no_root = '/'.join(path_parts) if path_parts else filename
        dir_levels = path_parts[:-1]
        # 修改日期
        last_mod = item.get('lastModified', '')
        if isinstance(last_mod, (int, float)):
            mod_time = datetime.fromtimestamp(last_mod / 1000).strftime('%Y-%m-%d')
        elif isinstance(last_mod, str) and last_mod:
            mod_time = last_mod[:10]
        else:
            mod_time = ''
        # 文件大小
        file_size_bytes = item.get('size', 0)
        formatted_size = format_file_size(file_size_bytes)

        file_info_list.append({
            'filename': filename,
            'rel_path': rel_path_no_root,
            'abs_path': rel_path,  # 远程模式无绝对路径，用原始路径占位
            'dir_levels': dir_levels,
            'mod_time': mod_time,
            'file_size': formatted_size
        })

    if not file_info_list:
        print("没有找到任何文件")
        return None

    create_excel(file_info_list, output_dir, 'file_index.xlsx')
    output_path = os.path.join(output_dir, 'file_index.xlsx')
    return output_path


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        workdir = sys.argv[1]
    else:
        workdir = os.path.dirname(__file__)
    build_index(workdir)
