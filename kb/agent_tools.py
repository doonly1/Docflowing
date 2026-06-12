import json
import logging
import os
import re
from typing import Any, Callable, Dict, List, Optional

logger = logging.getLogger(__name__)


MEMORY_SCHEMA = {
    "type": "function",
    "function": {
        "name": "memory",
        "description": (
            "Save durable information to persistent memory that survives across sessions. "
            "Memory is injected into future turns, so keep it compact and focused on facts "
            "that will still matter later.\n\n"
            "WHEN TO SAVE (do this proactively, don't wait to be asked):\n"
            "- User corrects you or says 'remember this' / 'don't do that again'\n"
            "- User shares a preference, habit, or personal detail (name, role, timezone, coding style)\n"
            "- You discover something about the environment (OS, installed tools, project structure)\n"
            "- You learn a convention, API quirk, or workflow specific to this user's setup\n"
            "- You identify a stable fact that will be useful again in future sessions\n\n"
            "PRIORITY: User preferences and corrections > environment facts > procedural knowledge. "
            "The most valuable memory prevents the user from having to repeat themselves.\n\n"
            "Do NOT save task progress, session outcomes, completed-work logs, or temporary TODO "
            "state to memory; use session_search to recall those from past transcripts.\n"
            "If you've discovered a new way to do something, solved a problem that could be "
            "necessary later, save it as a skill with the skill_manage tool.\n\n"
            "TWO TARGETS:\n"
            "- 'user': who the user is -- name, role, preferences, communication style, pet peeves\n"
            "- 'memory': your notes -- environment facts, project conventions, tool quirks, lessons learned\n\n"
            "ACTIONS: add (new entry), replace (update existing -- old_text identifies it), "
            "remove (delete -- old_text identifies it).\n\n"
            "SKIP: trivial/obvious info, things easily re-discovered, raw data dumps, and temporary task state."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["add", "replace", "remove"],
                    "description": "The action to perform."
                },
                "target": {
                    "type": "string",
                    "enum": ["memory", "user"],
                    "description": "Which memory store: 'memory' for personal notes, 'user' for user profile."
                },
                "content": {
                    "type": "string",
                    "description": "The entry content. Required for 'add' and 'replace'."
                },
                "old_text": {
                    "type": "string",
                    "description": "Short unique substring identifying the entry to replace or remove."
                }
            },
            "required": ["action", "target"]
        }
    }
}


SKILL_MANAGE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "skill_manage",
        "description": (
            "Manage the skill library -- create, view, patch, or list skills. "
            "Skills are reusable knowledge units that persist across sessions.\n\n"
            "WHEN TO CREATE A SKILL:\n"
            "- You solved a problem in a way that could be useful again\n"
            "- You developed a reusable procedure or workflow\n"
            "- You discovered a pattern worth remembering for future sessions\n\n"
            "ACTIONS:\n"
            "- create: Make a new skill with name, content, and optional category\n"
            "- patch: Update part of an existing skill (provide old_string and new_string)\n"
            "- view: Read a skill's content\n"
            "- list: List all skills or filter by category/state\n\n"
            "Do NOT create skills for one-time tasks, temporary state, or information that belongs in memory."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["create", "patch", "view", "list"],
                    "description": "The action to perform."
                },
                "name": {
                    "type": "string",
                    "description": "Skill name. Required for create, patch, view."
                },
                "content": {
                    "type": "string",
                    "description": "Skill content (markdown). Required for create."
                },
                "category": {
                    "type": "string",
                    "description": "Optional category for the skill."
                },
                "old_string": {
                    "type": "string",
                    "description": "Text to find in the skill. Required for patch."
                },
                "new_string": {
                    "type": "string",
                    "description": "Replacement text. Required for patch."
                },
                "state": {
                    "type": "string",
                    "enum": ["active", "stale", "archived"],
                    "description": "Filter by state for list action."
                }
            },
            "required": ["action"]
        }
    }
}


SESSION_SEARCH_SCHEMA = {
    "type": "function",
    "function": {
        "name": "session_search",
        "description": (
            "Search past conversation sessions for relevant information. "
            "Use this when you need to recall what happened in earlier conversations, "
            "find previous solutions to similar problems, or look up past decisions.\n\n"
            "Returns matching messages from past sessions with timestamps."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "Search query -- keywords or phrases to find in past sessions."
                },
                "limit": {
                    "type": "integer",
                    "description": "Maximum number of results to return. Default: 10."
                }
            },
            "required": ["query"]
        }
    }
}


WEB_SEARCH_SCHEMA = {
    "type": "function",
    "function": {
        "name": "web_search",
        "description": (
            "Search the web for current information. "
            "Use this when you need up-to-date information that may not be in the knowledge base, "
            "such as latest news, recent events, current API documentation, or real-time data.\n\n"
            "Returns a list of search results with titles, snippets, and URLs."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "The search query."
                },
                "max_results": {
                    "type": "integer",
                    "description": "Maximum number of results to return. Default: 5, max: 10."
                }
            },
            "required": ["query"]
        }
    }
}


WIKI_READ_SCHEMA = {
    "type": "function",
    "function": {
        "name": "wiki_read",
        "description": (
            "Read the full content of a knowledge base (wiki) document by its path. "
            "Use this when you've retrieved matched excerpts via the reference information "
            "but need to read an entire document for complete context.\n\n"
            "The path corresponds to the 'path' field in the sources list provided alongside "
            "search results. Only paths appearing in the sources are valid."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "The relative path of the wiki document to read, as shown in the sources list."
                }
            },
            "required": ["path"]
        }
    }
}


WIKI_SEARCH_SCHEMA = {
    "type": "function",
    "function": {
        "name": "wiki_search",
        "description": (
            "PRIMARY search tool: Search the user's local knowledge base (wiki) for documents matching a query. "
            "Always use this tool FIRST before falling back to web_search, because the wiki contains "
            "private/custom knowledge specific to the user and is faster than web search.\n\n"
            "Returns a list of matching documents with their paths, titles, and content snippets. "
            "Use wiki_read afterwards to read a specific document in full."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "The search query — keywords or phrases to find in the knowledge base."
                },
                "limit": {
                    "type": "integer",
                    "description": "Maximum number of results to return. Default: 5, max: 10."
                }
            },
            "required": ["query"]
        }
    }
}


# ==================== 文件库工具 ====================

FB_LIST_SCHEMA = {
    "type": "function",
    "function": {
        "name": "fb_list",
        "description": (
            "List all filebases (file libraries) accessible to the agent. "
            "Returns each filebase with its id, name, type, file count, and whether agent access is enabled. "
            "Use this to discover which filebases are available before browsing their contents."
        ),
        "parameters": {
            "type": "object",
            "properties": {},
            "required": []
        }
    }
}

FB_BROWSE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "fb_browse",
        "description": (
            "Browse the directory tree or file listing of a filebase. "
            "Returns files and subdirectories in the specified path. "
            "Use 'subdir' to navigate into subdirectories. Leave 'subdir' empty to browse the root."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "fb_id": {
                    "type": "string",
                    "description": "The filebase id (from fb_list)."
                },
                "subdir": {
                    "type": "string",
                    "description": "Relative path within the filebase (e.g. 'documents/subdir'). Empty string for root."
                }
            },
            "required": ["fb_id"]
        }
    }
}

FB_READ_SCHEMA = {
    "type": "function",
    "function": {
        "name": "fb_read",
        "description": (
            "Read the content of a text-based file (markdown, txt, source code, JSON, CSV, etc.) "
            "in a filebase. Also supports reading .docx and .xlsx files as plain text. "
            "Binary files like images, PDFs, and videos are not supported. "
            "Returns the file content and its detected type."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "fb_id": {
                    "type": "string",
                    "description": "The filebase id (from fb_list)."
                },
                "path": {
                    "type": "string",
                    "description": "Relative path of the file within the filebase (e.g. 'notes/meeting.md')."
                }
            },
            "required": ["fb_id", "path"]
        }
    }
}

FB_SEARCH_SCHEMA = {
    "type": "function",
    "function": {
        "name": "fb_search",
        "description": (
            "Search across all accessible filebases for documents matching keywords. "
            "Matches both file names and file content (text-based files only). "
            "Returns results grouped by filebase with file paths and match type (filename/content). "
            "Use this when you need to find specific information across filebases."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "Space-separated keywords to search for."
                },
                "limit": {
                    "type": "integer",
                    "description": "Maximum number of results per filebase. Default: 10."
                }
            },
            "required": ["query"]
        }
    }
}

FB_CREATE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "fb_create",
        "description": (
            "Create a new file or directory in a filebase. "
            "Files can be created with optional initial content. "
            "Subdirectories are supported via the 'parent' parameter. "
            "If no extension is provided for a file, '.md' is used by default. "
            "Cannot be used to create Office documents (.docx/.xlsx/.pptx). "
            "Duplicates will be auto-renamed with a numeric suffix."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "fb_id": {
                    "type": "string",
                    "description": "The filebase id (from fb_list)."
                },
                "type": {
                    "type": "string",
                    "enum": ["file", "dir"],
                    "description": "'file' to create a text file, 'dir' to create a directory."
                },
                "name": {
                    "type": "string",
                    "description": "Name of the file or directory to create."
                },
                "parent": {
                    "type": "string",
                    "description": "Parent directory path within the filebase. Empty string for root."
                },
                "content": {
                    "type": "string",
                    "description": "Initial file content (only for type='file')."
                }
            },
            "required": ["fb_id", "type", "name"]
        }
    }
}

FB_MOVE_RENAME_SCHEMA = {
    "type": "function",
    "function": {
        "name": "fb_move_rename",
        "description": (
            "Move one or more files/directories within a filebase to another directory, "
            "or rename a single file/directory. "
            "For moves: provide 'sources' (list of paths) and 'dest' (target directory path). "
            "For renames: provide 'path' and 'new_name' (name only, not full path). "
            "Cannot move items outside the filebase."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "fb_id": {
                    "type": "string",
                    "description": "The filebase id (from fb_list)."
                },
                "action": {
                    "type": "string",
                    "enum": ["move", "rename"],
                    "description": "'move' to move items to another directory, 'rename' to rename a single item."
                },
                "sources": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of relative paths to move (required for action='move')."
                },
                "dest": {
                    "type": "string",
                    "description": "Target directory path (required for action='move')."
                },
                "path": {
                    "type": "string",
                    "description": "Relative path of the item to rename (required for action='rename')."
                },
                "new_name": {
                    "type": "string",
                    "description": "New name for the item (required for action='rename')."
                }
            },
            "required": ["fb_id", "action"]
        }
    }
}


# ==================== 工具创建工具 ====================

TOOL_CREATE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "tool_create",
        "description": (
            "Create a new reusable tool with a custom function. "
            "The tool will be saved to disk and become available in future conversations. "
            "You define the tool's name, description, input JSON schema, and the Python code that implements it.\n\n"
            "WHEN TO CREATE A TOOL:\n"
            "- The user asks you to make a new tool for a recurring task\n"
            "- You need a capability that doesn't exist yet in your toolset\n"
            "- You want to automate a multi-step workflow\n\n"
            "HOW IT WORKS:\n"
            "- Provide a name (lowercase, underscores), description, and parameters schema\n"
            "- Write the execute() function body as Python code\n"
            "- The function receives 'args' (dict from LLM call) and 'user_id' (str)\n"
            "- Must return a JSON string via json.dumps()\n"
            "- The tool will be loaded and immediately available\n\n"
            "EXAMPLE execute_body:\n"
            '''```python
    value = args.get("value", 0)
    result = value * 2
    return json.dumps({"success": True, "result": result}, ensure_ascii=False)
    ```'''
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "name": {
                    "type": "string",
                    "description": "Tool name, lowercase with underscores (e.g. 'my_calculator'). Must be unique."
                },
                "description": {
                    "type": "string",
                    "description": "Clear description of what this tool does, when to use it."
                },
                "parameters_schema": {
                    "type": "object",
                    "description": "JSON Schema for the tool's parameters (the 'properties' field of the function parameters object). Each property must have type and description."
                },
                "required_params": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of required parameter names."
                },
                "execute_body": {
                    "type": "string",
                    "description": (
                        "Python code for the execute() function body. "
                        "Available variables: args (dict), user_id (str). "
                        "Must call json.dumps() and return the result. "
                        "Indentation must be exactly 4 spaces per level."
                    )
                }
            },
            "required": ["name", "description", "parameters_schema", "execute_body"]
        }
    }
}


TOOL_APPROVE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "tool_approve",
        "description": (
            "审批通过一个用户自建工具，允许其执行。"
            "仅当用户明确表示同意/允许执行某工具后，才调用此函数。"
            "调用前必须先向用户展示工具名和功能描述，获得明确许可。"
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "name": {
                    "type": "string",
                    "description": "用户自建工具的名称（创建时指定的 name）"
                }
            },
            "required": ["name"]
        }
    }
}


def _load_user_tools():
    """加载用户自建工具并更新全局列表"""
    try:
        from kb.user_tools import load_user_tools
        return load_user_tools()
    except Exception as e:
        logger.warning("无法加载用户工具: %s", e)
        return [], {}


# 加载用户自建工具
_USER_TOOL_SCHEMAS, _USER_TOOL_EXECUTORS = _load_user_tools()


ALL_TOOL_SCHEMAS = [
    MEMORY_SCHEMA, SKILL_MANAGE_SCHEMA, SESSION_SEARCH_SCHEMA,
    WEB_SEARCH_SCHEMA, WIKI_READ_SCHEMA, WIKI_SEARCH_SCHEMA,
    FB_LIST_SCHEMA, FB_BROWSE_SCHEMA, FB_READ_SCHEMA,
    FB_SEARCH_SCHEMA, FB_CREATE_SCHEMA, FB_MOVE_RENAME_SCHEMA,
    TOOL_CREATE_SCHEMA, TOOL_APPROVE_SCHEMA,
] + _USER_TOOL_SCHEMAS


def _check_fb_agent_allowed(fb_id: str) -> tuple:
    """检查文件库是否允许 agent 访问。返回 (allowed, error_json)"""
    try:
        from fb.database import get_db
        db = get_db()
        row = db.execute(
            "SELECT id, name, local_path, fb_agent_enabled, filebase_type FROM filebases WHERE id = ?",
            (fb_id,)
        ).fetchone()
        if not row:
            return False, json.dumps({"success": False, "error": f"文件库不存在: {fb_id}"}, ensure_ascii=False)
        # 远程文件库暂时不支持 agent 访问
        if row['filebase_type'] in ('remote',):
            return False, json.dumps({"success": False, "error": "远程文件库不支持 agent 访问"}, ensure_ascii=False)
        # NULL 视为默认开启（与 routes_base.py agent_settings GET 和前端逻辑一致）
        agent_enabled = row['fb_agent_enabled']
        if agent_enabled == 0:
            return False, json.dumps({"success": False, "error": f"文件库 '{row['name']}' 未允许 agent 访问"}, ensure_ascii=False)
        return True, None
    except Exception as e:
        return False, json.dumps({"success": False, "error": f"检查权限失败: {str(e)}"}, ensure_ascii=False)


def _execute_fb_list(args: Dict[str, Any], user_id: str) -> str:
    """列出所有允许 agent 访问的文件库"""
    try:
        from fb.database import get_db
        db = get_db()
        # agent 可见所有显式开启或未设置过开关的文件库（NULL 视为默认开启，与 _check_fb_agent_allowed 一致）
        rows = db.execute(
            "SELECT id, name, filebase_type, local_path, fb_agent_enabled, owner_id "
            "FROM filebases WHERE COALESCE(status, 'active') != 'trashed' AND (fb_agent_enabled IS NULL OR fb_agent_enabled != 0)"
        ).fetchall()
        results = []
        for row in rows:
            file_count = 0
            if row['local_path']:
                # 优先使用 worker 缓存 / sync_state，避免每次全量 os.walk
                try:
                    from fb.routes_base import _get_fb_file_count
                    file_count = _get_fb_file_count(row['id'], row['owner_id'])
                except Exception:
                    pass
            results.append({
                'id': row['id'],
                'name': row['name'],
                'type': row['filebase_type'] or 'local',
                'owner_id': row['owner_id'][:8] if row['owner_id'] else '',
                'file_count': file_count,
            })
        return json.dumps({"success": True, "filebases": results, "count": len(results)}, ensure_ascii=False)
    except Exception as e:
        logger.error("fb_list failed: %s", e)
        return json.dumps({"success": False, "error": f"列出文件库失败: {str(e)}"}, ensure_ascii=False)


def _execute_fb_browse(args: Dict[str, Any], user_id: str) -> str:
    """浏览文件库目录"""
    fb_id = args.get("fb_id", "")
    subdir = args.get("subdir", "")
    if not fb_id:
        return json.dumps({"success": False, "error": "fb_id is required"}, ensure_ascii=False)

    allowed, err = _check_fb_agent_allowed(fb_id)
    if not allowed:
        return err

    try:
        from fb.database import get_db
        db = get_db()
        row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (fb_id,)).fetchone()
        if not row or not row['local_path']:
            return json.dumps({"success": False, "error": "文件库路径不存在"}, ensure_ascii=False)
        local_path = row['local_path']
        target = os.path.join(local_path, subdir) if subdir else local_path
        target = os.path.normpath(target)
        if not target.startswith(os.path.normpath(local_path)):
            return json.dumps({"success": False, "error": "路径非法"}, ensure_ascii=False)
        if not os.path.isdir(target):
            return json.dumps({"success": False, "error": f"目录不存在: {subdir or '/'}"}, ensure_ascii=False)

        files = []
        dirs = []
        for entry in os.scandir(target):
            if entry.name.startswith('~$'):
                continue
            stat = entry.stat()
            rel = os.path.relpath(entry.path, local_path).replace('\\', '/')
            if entry.is_dir():
                dirs.append({'name': entry.name, 'path': rel})
            elif entry.is_file():
                files.append({
                    'name': entry.name,
                    'path': rel,
                    'size': stat.st_size,
                    'mtime': stat.st_mtime,
                })
        dirs.sort(key=lambda x: x['name'].lower())
        files.sort(key=lambda x: x['name'].lower())

        return json.dumps({
            "success": True,
            "current_path": subdir.replace('\\', '/') if subdir else '',
            "directories": dirs,
            "files": files,
            "total_dirs": len(dirs),
            "total_files": len(files),
        }, ensure_ascii=False)
    except Exception as e:
        logger.error("fb_browse failed: %s", e)
        return json.dumps({"success": False, "error": f"浏览文件库失败: {str(e)}"}, ensure_ascii=False)


def _execute_fb_read(args: Dict[str, Any], user_id: str) -> str:
    """读取文件内容"""
    fb_id = args.get("fb_id", "")
    path = args.get("path", "")
    if not fb_id or not path:
        return json.dumps({"success": False, "error": "fb_id and path are required"}, ensure_ascii=False)

    allowed, err = _check_fb_agent_allowed(fb_id)
    if not allowed:
        return err

    try:
        from fb.database import get_db
        db = get_db()
        row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (fb_id,)).fetchone()
        if not row or not row['local_path']:
            return json.dumps({"success": False, "error": "文件库路径不存在"}, ensure_ascii=False)
        local_path = row['local_path']
        file_path = os.path.normpath(os.path.join(local_path, path))
        if not file_path.startswith(os.path.normpath(local_path)):
            return json.dumps({"success": False, "error": "路径非法"}, ensure_ascii=False)
        if not os.path.isfile(file_path):
            return json.dumps({"success": False, "error": f"文件不存在: {path}"}, ensure_ascii=False)

        ext = os.path.splitext(file_path)[1].lower()

        # 文本文件
        text_exts = {'.md', '.txt', '.html', '.htm', '.xml', '.json', '.csv', '.yaml', '.yml',
                     '.py', '.js', '.css', '.ts', '.tsx', '.jsx', '.sh', '.bat', '.conf', '.ini',
                     '.cfg', '.env', '.log', '.sql', '.rb', '.go', '.rs', '.java', '.c', '.cpp',
                     '.h', '.hpp', '.toml', '.lock', '.gradle', '.m', '.swift', '.kt', '.scala'}
        if ext in text_exts:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                if len(content) > 50000:
                    content = content[:50000] + "\n\n... (truncated at 50000 characters)"
                return json.dumps({"success": True, "content": content, "file_type": ext}, ensure_ascii=False)
            except UnicodeDecodeError:
                try:
                    with open(file_path, 'r', encoding='gbk') as f:
                        content = f.read()
                    return json.dumps({"success": True, "content": content, "file_type": ext}, ensure_ascii=False)
                except Exception:
                    return json.dumps({"success": False, "error": "无法以文本方式读取此文件"}, ensure_ascii=False)

        # .docx
        if ext == '.docx':
            try:
                from docx import Document
                doc = Document(file_path)
                text = '\n'.join(p.text for p in doc.paragraphs)
                if len(text) > 50000:
                    text = text[:50000] + "\n\n... (truncated at 50000 characters)"
                return json.dumps({"success": True, "content": text, "file_type": ext}, ensure_ascii=False)
            except Exception:
                return json.dumps({"success": False, "error": "无法读取 docx 内容"}, ensure_ascii=False)

        # .xlsx
        if ext in ('.xlsx', '.xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                lines = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    lines.append(f'# {sheet_name}')
                    for row_data in ws.iter_rows(values_only=True):
                        lines.append('\t'.join(str(c) if c is not None else '' for c in row_data))
                text = '\n'.join(lines)
                if len(text) > 50000:
                    text = text[:50000] + "\n\n... (truncated at 50000 characters)"
                return json.dumps({"success": True, "content": text, "file_type": ext}, ensure_ascii=False)
            except Exception:
                return json.dumps({"success": False, "error": "无法读取 xlsx 内容"}, ensure_ascii=False)

        return json.dumps({"success": False, "error": f"不支持在线读取 {ext} 类型文件"}, ensure_ascii=False)
    except Exception as e:
        logger.error("fb_read failed: %s", e)
        return json.dumps({"success": False, "error": f"读取文件失败: {str(e)}"}, ensure_ascii=False)


def _execute_fb_search(args: Dict[str, Any], user_id: str) -> str:
    """搜索文件"""
    query = args.get("query", "")
    limit = args.get("limit", 10)
    if not query:
        return json.dumps({"success": False, "error": "query is required"}, ensure_ascii=False)

    try:
        from fb.database import get_db
        from fb.routes_search import _search_local_dir

        db = get_db()
        # 只搜索允许 agent 访问的文件库
        rows = db.execute(
            "SELECT id, name, local_path FROM filebases "
            "WHERE COALESCE(status, 'active') != 'trashed' AND (fb_agent_enabled IS NULL OR fb_agent_enabled != 0) "
            "AND local_path != '' AND filebase_type != 'remote'"
        ).fetchall()

        keywords = query.lower().split()
        all_results = []
        for row in rows:
            if row['local_path'] and os.path.isdir(row['local_path']):
                fb_results = _search_local_dir(row['local_path'], row['id'], row['name'], keywords)
                all_results.extend(fb_results)

        # 限制总数
        if limit and len(all_results) > limit:
            all_results = all_results[:limit]

        return json.dumps({
            "success": True,
            "results": all_results,
            "count": len(all_results),
            "query": query,
        }, ensure_ascii=False)
    except Exception as e:
        logger.error("fb_search failed: %s", e)
        return json.dumps({"success": False, "error": f"搜索失败: {str(e)}"}, ensure_ascii=False)


def _execute_fb_create(args: Dict[str, Any], user_id: str) -> str:
    """创建文件或目录"""
    fb_id = args.get("fb_id", "")
    create_type = args.get("type", "file")
    name = args.get("name", "")
    parent = args.get("parent", "")
    content = args.get("content", "")

    if not fb_id or not name:
        return json.dumps({"success": False, "error": "fb_id and name are required"}, ensure_ascii=False)

    allowed, err = _check_fb_agent_allowed(fb_id)
    if not allowed:
        return err

    if '/' in name or '\\' in name:
        return json.dumps({"success": False, "error": "名称不能包含路径分隔符"}, ensure_ascii=False)

    try:
        from fb.database import get_db
        db = get_db()
        row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (fb_id,)).fetchone()
        if not row or not row['local_path']:
            return json.dumps({"success": False, "error": "文件库路径不存在"}, ensure_ascii=False)
        local_path = row['local_path']

        target_dir = os.path.join(local_path, parent) if parent else local_path
        target_dir = os.path.normpath(target_dir)
        if not target_dir.startswith(os.path.normpath(local_path)):
            return json.dumps({"success": False, "error": "路径非法"}, ensure_ascii=False)

        if create_type == 'dir':
            new_dir = os.path.join(target_dir, name)
            counter = 1
            orig = name
            while os.path.exists(new_dir) and counter < 100:
                name = f'{orig}_{counter}'
                new_dir = os.path.join(target_dir, name)
                counter += 1
            if os.path.exists(new_dir):
                return json.dumps({"success": False, "error": "无法生成唯一的目录名称"}, ensure_ascii=False)
            os.makedirs(new_dir, exist_ok=True)
            rel = os.path.relpath(new_dir, local_path).replace('\\', '/')
            return json.dumps({"success": True, "type": "dir", "path": rel}, ensure_ascii=False)

        # 创建文件
        base, ext = os.path.splitext(name)
        if ext:
            filename = name
            default_ext = ext
        else:
            filename = name + '.md'
            default_ext = '.md'
        file_path = os.path.join(target_dir, filename)

        counter = 1
        while os.path.exists(file_path) and counter < 100:
            filename = f'{base}_{counter}{default_ext}'
            file_path = os.path.join(target_dir, filename)
            counter += 1
        if os.path.exists(file_path):
            return json.dumps({"success": False, "error": "无法生成唯一的文件名"}, ensure_ascii=False)

        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content or '')

        # 触发同步（直接调 worker，不依赖 Flask 请求上下文 g）
        try:
            from kb.sync_worker import get_sync_worker
            worker = get_sync_worker()
            worker.adjust_file_count(user_id, fb_id, 1)
            worker._trigger_sync(user_id, fb_id)
        except Exception:
            logger.exception("Failed to trigger sync after fb_create")

        rel = os.path.relpath(file_path, local_path).replace('\\', '/')
        return json.dumps({"success": True, "type": "file", "path": rel, "filename": os.path.basename(file_path)}, ensure_ascii=False)
    except Exception as e:
        logger.error("fb_create failed: %s", e)
        return json.dumps({"success": False, "error": f"创建失败: {str(e)}"}, ensure_ascii=False)


def _execute_fb_move_rename(args: Dict[str, Any], user_id: str) -> str:
    """移动或重命名文件/目录"""
    fb_id = args.get("fb_id", "")
    action = args.get("action", "")

    if not fb_id or not action:
        return json.dumps({"success": False, "error": "fb_id and action are required"}, ensure_ascii=False)

    allowed, err = _check_fb_agent_allowed(fb_id)
    if not allowed:
        return err

    try:
        from fb.database import get_db
        db = get_db()
        row = db.execute("SELECT local_path FROM filebases WHERE id = ?", (fb_id,)).fetchone()
        if not row or not row['local_path']:
            return json.dumps({"success": False, "error": "文件库路径不存在"}, ensure_ascii=False)
        local_path = row['local_path']

        if action == 'move':
            sources = args.get("sources", [])
            dest = args.get("dest", "")
            if not sources or not dest:
                return json.dumps({"success": False, "error": "sources and dest are required for move"}, ensure_ascii=False)

            dest_path = os.path.normpath(os.path.join(local_path, dest))
            if not dest_path.startswith(os.path.normpath(local_path)):
                return json.dumps({"success": False, "error": "目标目录非法"}, ensure_ascii=False)
            os.makedirs(dest_path, exist_ok=True)

            moved = 0
            errors = []
            for src in sources:
                src_path = os.path.normpath(os.path.join(local_path, src))
                if not src_path.startswith(os.path.normpath(local_path)):
                    errors.append(f'{src}: 路径非法')
                    continue
                if not os.path.exists(src_path):
                    errors.append(f'{src}: 不存在')
                    continue
                target = os.path.join(dest_path, os.path.basename(src_path))
                if os.path.exists(target):
                    errors.append(f'{src}: 目标位置已存在同名项目')
                    continue
                import shutil
                shutil.move(src_path, target)
                moved += 1

            # 触发同步（直接调 worker，不依赖 Flask 请求上下文 g）
            try:
                from kb.sync_worker import get_sync_worker
                worker = get_sync_worker()
                worker._trigger_sync(user_id, fb_id)
            except Exception:
                logger.exception("Failed to trigger sync after fb_move")

            return json.dumps({"success": True, "moved": moved, "errors": errors}, ensure_ascii=False)

        elif action == 'rename':
            path = args.get("path", "")
            new_name = args.get("new_name", "")
            if not path or not new_name:
                return json.dumps({"success": False, "error": "path and new_name are required for rename"}, ensure_ascii=False)
            if '/' in new_name or '\\' in new_name:
                return json.dumps({"success": False, "error": "新名称不能包含路径分隔符"}, ensure_ascii=False)

            old = os.path.normpath(os.path.join(local_path, path))
            if not old.startswith(os.path.normpath(local_path)):
                return json.dumps({"success": False, "error": "路径非法"}, ensure_ascii=False)
            if not os.path.exists(old):
                return json.dumps({"success": False, "error": "文件或目录不存在"}, ensure_ascii=False)

            parent_dir = os.path.dirname(old)
            new_path = os.path.join(parent_dir, new_name)
            if os.path.exists(new_path):
                return json.dumps({"success": False, "error": "同名文件或目录已存在"}, ensure_ascii=False)

            os.rename(old, new_path)
            new_rel = os.path.relpath(new_path, local_path).replace('\\', '/')

            try:
                from kb.sync_worker import get_sync_worker
                worker = get_sync_worker()
                worker._trigger_sync(user_id, fb_id)
            except Exception:
                logger.exception("Failed to trigger sync after fb_rename")

            return json.dumps({"success": True, "new_path": new_rel}, ensure_ascii=False)
        else:
            return json.dumps({"success": False, "error": f"Unknown action: {action}"}, ensure_ascii=False)
    except Exception as e:
        logger.error("fb_move_rename failed: %s", e)
        return json.dumps({"success": False, "error": f"操作失败: {str(e)}"}, ensure_ascii=False)


def _reload_user_tools():
    """重新加载用户工具并更新全局变量"""
    global _USER_TOOL_SCHEMAS, _USER_TOOL_EXECUTORS, ALL_TOOL_SCHEMAS
    try:
        from kb.user_tools import reload_user_tools
        schemas, executors = reload_user_tools()
        _USER_TOOL_SCHEMAS = schemas
        _USER_TOOL_EXECUTORS = executors
        # 重建 ALL_TOOL_SCHEMAS
        ALL_TOOL_SCHEMAS = [
            MEMORY_SCHEMA, SKILL_MANAGE_SCHEMA, SESSION_SEARCH_SCHEMA,
            WEB_SEARCH_SCHEMA, WIKI_READ_SCHEMA, WIKI_SEARCH_SCHEMA,
            FB_LIST_SCHEMA, FB_BROWSE_SCHEMA, FB_READ_SCHEMA,
            FB_SEARCH_SCHEMA, FB_CREATE_SCHEMA, FB_MOVE_RENAME_SCHEMA,
            TOOL_CREATE_SCHEMA, TOOL_APPROVE_SCHEMA,
        ] + schemas
    except Exception as e:
        logger.error("重新加载用户工具失败: %s", e)


def _validate_tool_code_ast(execute_body: str) -> Optional[str]:
    """用 AST 解析工具函数体，校验安全性（白名单 + 黑名单双层）。

    规则：
    - import X：X 必须在 _ALLOWED_MODULES 中
    - from X import Y：X 的顶级模块不在 _BLOCKED_MODULES 中，
      或属于 _FROM_ONLY_ALLOWED 的特殊豁免
    - 禁止危险 builtins：__import__, open, exec, eval, compile, input
    - 禁止 pathlib 写方法：unlink, write_text, mkdir 等

    Returns: None 通过，str 错误信息
    """
    import ast

    _BLOCKED_MODULES = {
        'os', 'shutil', 'subprocess', 'socket', 'ctypes', 'importlib',
        'pickle', 'sys', 'http', 'urllib', 'ssl', 'smtplib', 'ftplib',
        'multiprocessing', 'threading', 'code', 'codeop',
        'tempfile', 'atexit', 'signal', 'platform', 'inspect',
        'dis', 'gc', 'sysconfig', 'pkgutil',
    }
    _FROM_ONLY_ALLOWED = {
        'os': {'path'},                              # from os import path
        'os.path': None,                              # from os.path import *
    }
    _ALLOWED_MODULES = {
        'json', 're', 'datetime', 'collections', 'typing', 'math',
        'itertools', 'functools', 'textwrap', 'string', 'numbers',
        'decimal', 'fractions', 'random', 'statistics',
        'hashlib', 'base64', 'binascii',
        'time', 'calendar', 'copy', 'enum', 'dataclasses',
        'csv', 'glob', 'fnmatch', 'difflib', 'uuid',
    }
    _PATHLIB_BLOCKED = {
        'unlink', 'rmdir', 'chmod', 'symlink_to', 'hardlink_to',
        'rename', 'replace', 'write_bytes', 'write_text', 'open',
        'mkdir', 'touch', 'lchmod',
    }
    _BLOCKED_BUILTINS = {'__import__', 'open', 'exec', 'eval', 'compile', 'input', 'breakpoint'}

    try:
        tree = ast.parse(execute_body)
    except SyntaxError as e:
        return f"语法错误: {e}"

    for node in ast.walk(tree):
        # --- import X ---
        if isinstance(node, ast.Import):
            for alias in node.names:
                name = alias.name
                top = name.split('.')[0]
                if top in _BLOCKED_MODULES:
                    return f"禁止导入模块: {name}"
                if name not in _ALLOWED_MODULES and top not in _ALLOWED_MODULES:
                    return f"不允许导入模块: {name}"

        # --- from X import Y ---
        if isinstance(node, ast.ImportFrom):
            mod = node.module or ''
            top = mod.split('.')[0]
            if top in _BLOCKED_MODULES:
                # 特殊豁免：from os import path
                special = _FROM_ONLY_ALLOWED.get(mod) or _FROM_ONLY_ALLOWED.get(top, set())
                if special is not None:
                    for alias in node.names:
                        if alias.name not in special:
                            return f"禁止从 {mod} 导入 {alias.name}"
                    continue
                return f"禁止导入模块: {mod}"
            # 非黑名单模块：顶级必须在白名单
            if top not in _ALLOWED_MODULES and mod not in _ALLOWED_MODULES:
                return f"不允许导入模块: {mod}"

        # --- builtins 引用 ---
        if isinstance(node, ast.Name) and node.id in _BLOCKED_BUILTINS:
            return f"禁止使用内置函数: {node.id}"

        # --- pathlib 危险方法调用 ---
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Attribute):
            if node.func.attr in _PATHLIB_BLOCKED:
                return f"禁止调用写方法: {node.func.attr}"

    return None


def _execute_tool_create(args: Dict[str, Any], user_id: str) -> str:
    """创建用户自建工具"""
    name = (args.get("name") or "").strip()
    description = (args.get("description") or "").strip()
    parameters_schema = args.get("parameters_schema", {})
    required_params = args.get("required_params", [])
    execute_body = (args.get("execute_body") or "").strip()

    if not name or not description or not execute_body:
        return json.dumps({"success": False, "error": "name, description and execute_body are required"}, ensure_ascii=False)

    # 校验名称
    if not re.match(r'^[a-z][a-z0-9_]*$', name):
        return json.dumps({"success": False, "error": "工具名必须是小写字母开头，只含小写字母、数字和下划线"}, ensure_ascii=False)

    # 不能覆盖内置工具
    builtin_tools = {
        "memory", "skill_manage", "session_search", "web_search",
        "wiki_read", "wiki_search",
        "fb_list", "fb_browse", "fb_read", "fb_search", "fb_create", "fb_move_rename",
        "tool_create",
    }
    if name in builtin_tools:
        return json.dumps({"success": False, "error": f"工具名 '{name}' 与内置工具冲突"}, ensure_ascii=False)
    if name in _USER_TOOL_EXECUTORS:
        return json.dumps({"success": False, "error": f"工具名 '{name}' 已存在，请先删除再重建"}, ensure_ascii=False)

    # 安全校验：AST 白名单解析（替代旧字符串黑名单，不可绕过）
    ast_error = _validate_tool_code_ast(execute_body)
    if ast_error:
        return json.dumps({"success": False, "error": f"代码安全校验不通过: {ast_error}"}, ensure_ascii=False)

    # 构建工具文件
    schema = {
        "type": "function",
        "function": {
            "name": name,
            "description": description,
            "parameters": {
                "type": "object",
                "properties": parameters_schema,
                "required": required_params
            }
        }
    }
    schema_json = json.dumps(schema, ensure_ascii=False, indent=2)

    # 缩进 execute_body（4 空格）
    indented_body = '\n'.join('    ' + line for line in execute_body.split('\n'))

    tool_code = f'''"""
User-created tool: {name}
Description: {description}
Created by: agent ({user_id[:16]})
"""

import json

SCHEMA = {schema_json}


def execute(args: dict, user_id: str) -> str:
{indented_body}
'''

    # 写出文件到运行时目录 user_tools/（由 server.workspace 统一解析）
    from server.workspace import _get_workspace_dir
    base_dir = _get_workspace_dir()
    tools_dir = os.path.join(base_dir, 'user_tools')
    os.makedirs(tools_dir, exist_ok=True)
    fpath = os.path.join(tools_dir, f'{name}.py')

    if os.path.exists(fpath):
        return json.dumps({"success": False, "error": f"文件 {name}.py 已存在"}, ensure_ascii=False)

    try:
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(tool_code)
    except Exception as e:
        return json.dumps({"success": False, "error": f"写入工具文件失败: {str(e)}"}, ensure_ascii=False)

    # 重新加载用户工具
    _reload_user_tools()

    # 验证加载成功
    if name not in _USER_TOOL_EXECUTORS:
        error_msg = f"工具 '{name}' 文件已创建但加载失败，请检查代码语法"
        return json.dumps({"success": False, "error": error_msg}, ensure_ascii=False)

    return json.dumps({
        "success": True,
        "message": f"工具 '{name}' 已创建，需要用户审批通过后才能执行",
        "tool_name": name,
        "pending_approval": True,
    }, ensure_ascii=False)


# ==================== 用户工具审批系统 ====================

_APPROVALS_FILE = None  # lazy init


def _get_approvals_path():
    global _APPROVALS_FILE
    if _APPROVALS_FILE is None:
        from server.workspace import _get_workspace_dir
        base = _get_workspace_dir()
        _APPROVALS_FILE = os.path.join(base, 'user_tools', '.approvals.json')
        os.makedirs(os.path.dirname(_APPROVALS_FILE), exist_ok=True)
    return _APPROVALS_FILE


def _load_approvals() -> dict:
    """加载审批状态文件"""
    path = _get_approvals_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _save_approvals(data: dict):
    """保存审批状态文件"""
    path = _get_approvals_path()
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        logger.exception("保存审批状态失败")


def _is_tool_approved(tool_name: str) -> bool:
    """检查用户工具是否已审批通过"""
    data = _load_approvals()
    entry = data.get(tool_name, {})
    return entry.get('approved', False)


def _approve_tool(tool_name: str, user_id: str):
    """审批通过用户工具"""
    data = _load_approvals()
    data[tool_name] = {'approved': True, 'approved_at': __import__('time').time(), 'user_id': user_id[:16]}
    _save_approvals(data)


def _reject_tool(tool_name: str):
    """拒绝用户工具"""
    data = _load_approvals()
    if tool_name in data:
        data[tool_name]['approved'] = False
    else:
        data[tool_name] = {'approved': False, 'rejected_at': __import__('time').time()}
    _save_approvals(data)


def execute_tool_call(tool_name: str, args: Dict[str, Any], user_id: str) -> str:
    try:
        if tool_name == "memory":
            return _execute_memory(args, user_id)
        elif tool_name == "skill_manage":
            return _execute_skill_manage(args, user_id)
        elif tool_name == "session_search":
            return _execute_session_search(args, user_id)
        elif tool_name == "web_search":
            return _execute_web_search(args)
        elif tool_name == "wiki_read":
            return _execute_wiki_read(args, user_id)
        elif tool_name == "wiki_search":
            return _execute_wiki_search(args, user_id)
        elif tool_name == "fb_list":
            return _execute_fb_list(args, user_id)
        elif tool_name == "fb_browse":
            return _execute_fb_browse(args, user_id)
        elif tool_name == "fb_read":
            return _execute_fb_read(args, user_id)
        elif tool_name == "fb_search":
            return _execute_fb_search(args, user_id)
        elif tool_name == "fb_create":
            return _execute_fb_create(args, user_id)
        elif tool_name == "fb_move_rename":
            return _execute_fb_move_rename(args, user_id)
        elif tool_name == "tool_create":
            return _execute_tool_create(args, user_id)
        elif tool_name == "tool_approve":
            tool_name_approve = (args.get("name") or "").strip()
            if not tool_name_approve:
                return json.dumps({"success": False, "error": "缺少工具名"})
            _approve_tool(tool_name_approve, user_id)
            return json.dumps({"success": True, "message": f"工具 '{tool_name_approve}' 已审批通过，现在可以执行了"})
        elif tool_name in _USER_TOOL_EXECUTORS:
            # 用户工具需要审批通过才能执行
            if not _is_tool_approved(tool_name):
                return json.dumps({
                    "success": False,
                    "error": f"工具 '{tool_name}' 需要审批通过后才能执行",
                    "require_approval": True,
                    "tool_name": tool_name,
                }, ensure_ascii=False)
            return _USER_TOOL_EXECUTORS[tool_name](args, user_id)
        else:
            return json.dumps({"success": False, "error": f"Unknown tool: {tool_name}"}, ensure_ascii=False)
    except Exception as e:
        logger.error("Tool execution failed: %s %s -> %s", tool_name, args, e)
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)


def _validate_content_safety(content: str) -> Optional[str]:
    from .file_safety import validate_content_safety
    return validate_content_safety(content)


def _execute_memory(args: Dict[str, Any], user_id: str) -> str:
    from .memory import get_memory_store

    store = get_memory_store(user_id)
    action = args.get("action", "")
    target = args.get("target", "memory")

    if action == "add":
        content = args.get("content", "")
        if not content:
            return json.dumps({"success": False, "error": "content is required for add"}, ensure_ascii=False)
        result = store.add(target, content)
    elif action == "replace":
        old_text = args.get("old_text", "")
        content = args.get("content", "")
        if not old_text or not content:
            return json.dumps({"success": False, "error": "old_text and content are required for replace"}, ensure_ascii=False)
        result = store.replace(target, old_text, content)
    elif action == "remove":
        old_text = args.get("old_text", "")
        if not old_text:
            return json.dumps({"success": False, "error": "old_text is required for remove"}, ensure_ascii=False)
        result = store.remove(target, old_text)
    else:
        result = {"success": False, "error": f"Unknown action: {action}"}

    return json.dumps(result, ensure_ascii=False)


def _execute_skill_manage(args: Dict[str, Any], user_id: str) -> str:
    from .skills.manager import create_skill, patch_skill, get_skill, list_skills
    from .file_safety import validate_skill_path

    action = args.get("action", "")

    if action == "create":
        name = args.get("name", "")
        content = args.get("content", "")
        category = args.get("category")
        if not name or not content:
            return json.dumps({"success": False, "error": "name and content are required"}, ensure_ascii=False)
        path_error = validate_skill_path(user_id, name, None)
        if path_error:
            return json.dumps({"success": False, "error": path_error}, ensure_ascii=False)
        content_error = _validate_content_safety(content)
        if content_error:
            return json.dumps({"success": False, "error": content_error}, ensure_ascii=False)
        result = create_skill(user_id, name, content, category=category, created_by="agent")
    elif action == "patch":
        name = args.get("name", "")
        old_string = args.get("old_string", "")
        new_string = args.get("new_string", "")
        if not name or not old_string or not new_string:
            return json.dumps({"success": False, "error": "name, old_string and new_string are required"}, ensure_ascii=False)
        content_error = _validate_content_safety(new_string)
        if content_error:
            return json.dumps({"success": False, "error": content_error}, ensure_ascii=False)
        result = patch_skill(user_id, name, old_string, new_string)
    elif action == "view":
        name = args.get("name", "")
        if not name:
            return json.dumps({"success": False, "error": "name is required"}, ensure_ascii=False)
        result = get_skill(user_id, name)
    elif action == "list":
        category = args.get("category")
        state = args.get("state")
        skills = list_skills(user_id, category=category, state=state)
        result = {"success": True, "skills": skills, "count": len(skills)}
    else:
        result = {"success": False, "error": f"Unknown action: {action}"}

    return json.dumps(result, ensure_ascii=False)


def _execute_session_search(args: Dict[str, Any], user_id: str) -> str:
    from .session_db import get_session_db

    query = args.get("query", "")
    limit = args.get("limit", 10)
    if not query:
        return json.dumps({"success": False, "error": "query is required"}, ensure_ascii=False)

    db = get_session_db(user_id)
    results = db.search_messages(query, user_id=user_id, limit=limit)
    return json.dumps({"success": True, "results": results, "count": len(results)}, ensure_ascii=False)


def _execute_web_search(args: Dict[str, Any]) -> str:
    query = args.get("query", "")
    max_results = min(args.get("max_results", 5), 10)
    if not query:
        return json.dumps({"success": False, "error": "query is required"}, ensure_ascii=False)

    try:
        import requests
        from bs4 import BeautifulSoup

        resp = requests.post(
            "https://html.duckduckgo.com/html/",
            data={"q": query},
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            },
            timeout=15,
        )
        resp.raise_for_status()

        soup = BeautifulSoup(resp.text, "html.parser")
        results = []
        for item in soup.select(".result")[:max_results]:
            title_el = item.select_one(".result__title a")
            snippet_el = item.select_one(".result__snippet")
            if not title_el:
                continue
            results.append({
                "title": title_el.get_text(strip=True),
                "url": title_el.get("href", ""),
                "snippet": snippet_el.get_text(strip=True) if snippet_el else "",
            })

        return json.dumps({
            "success": True,
            "results": results,
            "count": len(results),
            "query": query,
        }, ensure_ascii=False)

    except ImportError:
        return json.dumps({
            "success": False,
            "error": "Web search requires beautifulsoup4. Install with: pip install beautifulsoup4",
        }, ensure_ascii=False)
    except requests.exceptions.Timeout:
        return json.dumps({"success": False, "error": "Web search timed out"}, ensure_ascii=False)
    except Exception as e:
        logger.error("Web search failed: %s", e)
        return json.dumps({"success": False, "error": f"Web search failed: {str(e)}"}, ensure_ascii=False)


def _extract_from_keyword(text: str, keywords: list) -> str:
    """从关键词所在句子的开头截取文本，使上下文更紧凑且有针对性。"""
    pos = -1
    text_lower = text.lower()
    for kw in keywords:
        p = text_lower.find(kw.lower())
        if p != -1:
            pos = p
            break
    if pos == -1:
        return text
    sentence_start = 0
    for sep in ('。', '！', '？', '!', '?', '\n'):
        idx = text.rfind(sep, 0, pos)
        if idx != -1:
            sentence_start = max(sentence_start, idx + len(sep))
    return text[sentence_start:].strip()


def _extract_relevant_snippets(content: str, keywords: list, max_chars: int = 3000) -> str:
    """提取包含关键词的段落，从关键词所在句子开头截取，限制总长度。

    Args:
        content: 文档全文
        keywords: 关键词列表（小写）
        max_chars: 最大返回字符数

    Returns:
        经截取和修剪后的文本片段
    """
    if not content.strip():
        return ""

    # 按空行分割段落，提取包含关键词的段落
    paragraphs = re.split(r'\n\s*\n', content)
    matched_bodies = []
    for para in paragraphs:
        if any(kw in para.lower() for kw in keywords):
            matched_bodies.append(para.strip())

    # 兜底：无匹配段落则展示开头
    if not matched_bodies:
        matched_bodies = [content.strip()[:300]]

    # 从关键词所在句子开头截取
    trimmed_bodies = []
    for para in matched_bodies:
        trimmed_bodies.append(_extract_from_keyword(para, keywords))

    # 限制总长
    snippet = '\n'.join(trimmed_bodies)
    if len(snippet) > max_chars:
        snippet = snippet[:max_chars] + "..."

    return snippet


def _execute_wiki_read(args: Dict[str, Any], user_id: str) -> str:
    path = args.get("path", "")
    if not path:
        return json.dumps({"success": False, "error": "path is required"}, ensure_ascii=False)

    from .database import get_db
    conn = get_db(user_id)
    row = conn.execute(
        "SELECT content FROM wiki_fts WHERE usr_id = ? AND path = ?",
        (user_id, path)
    ).fetchone()
    if not row:
        return json.dumps({"success": False, "error": f"Document not found: {path}"}, ensure_ascii=False)

    content = row['content']
    if len(content) > 50000:
        content = content[:50000] + "\n\n... (truncated at 50000 characters)"
    return json.dumps({"success": True, "path": path, "content": content}, ensure_ascii=False)


def _execute_wiki_search(args: Dict[str, Any], user_id: str) -> str:
    from .search import search_wiki
    from .database import get_db

    query = args.get("query", "")
    limit = min(args.get("limit", 5), 10)
    if not query:
        return json.dumps({"success": False, "error": "query is required"}, ensure_ascii=False)

    try:
        results = search_wiki(user_id, query)
        if limit:
            results = results[:limit]

        conn = get_db(user_id)
        enriched = []
        keywords = query.lower().split()
        max_snippet_chars = 3000

        for r in results:
            row = conn.execute(
                "SELECT content FROM wiki_fts WHERE usr_id = ? AND path = ?",
                (user_id, r['path'])
            ).fetchone()
            content = row['content'] if row else ""
            if not content.strip():
                continue

            snippet = _extract_relevant_snippets(content, keywords, max_snippet_chars)
            enriched.append({
                "path": r['path'],
                "title": r['title'],
                "content_snippet": snippet,
            })

        return json.dumps({
            "success": True,
            "results": enriched,
            "count": len(enriched),
            "query": query,
        }, ensure_ascii=False)
    except Exception as e:
        logger.error("Wiki search failed: %s", e)
        return json.dumps({"success": False, "error": f"Wiki search failed: {str(e)}"}, ensure_ascii=False)
